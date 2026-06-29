"""
pipeline_server.py — FastAPI HTTP server wrapping the chroma pipeline scripts.
n8n calls these endpoints instead of using Execute Command (unreliable on Windows).

Endpoints:
  GET  /health              → liveness check
  POST /run/parse_cdf       → runs parse_cdf.py on a .cdf file → returns peaks JSON
  POST /run/enrich          → calls T2 LiteLLM to enrich compound names
  GET  /results/{sample_id} → fetch a previously processed result

Start: python scripts/pipeline_server.py
Port:  8001
"""
import json
from contextlib import asynccontextmanager
import os
import subprocess
import sys
import threading
import time
from pathlib import Path

import requests
import uvicorn
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, UploadFile, File, Depends, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from pydantic import BaseModel

load_dotenv(dotenv_path=Path(__file__).parent.parent / ".env")

BASE_DIR    = Path(__file__).parent.parent
RAW_DIR     = BASE_DIR / "raw_data"
PROC_DIR    = BASE_DIR / "processed" / ".cache"
RESULTS_DIR = BASE_DIR / "processed_results"
SCRIPTS_DIR = BASE_DIR / "scripts"
PYTHON      = sys.executable
LITELLM_URL = "http://127.0.0.1:4000"

WATCHER_STATUS = {
    "active": False,
    "current_file": None,
    "stage": "idle",       # "idle", "parsing", "deconvolving", "matching", "enriching", "generating_excel", "complete", "failed"
    "stages_completed": [], # ["parse", "deconv", "match", "enrich"]
    "last_error": None
}

PROC_DIR.mkdir(parents=True, exist_ok=True)
RESULTS_DIR.mkdir(exist_ok=True)

@asynccontextmanager
async def lifespan(app: FastAPI):
    start_folder_watcher()
    yield

app = FastAPI(
    title="CHROMA-AGENT-ALPHA Pipeline Server",
    description="HTTP API wrapping GC-MS chromatography pipeline stages for n8n orchestration.",
    version="1.0.0",
    lifespan=lifespan,
)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


# ── Security Utilities & Authentication ──────────────────────────────

security = HTTPBasic()

def authenticate_user(credentials: HTTPBasicCredentials = Depends(security)):
    api_user = os.environ.get("CHROMA_API_USER", "tester")
    api_password = os.environ.get("CHROMA_API_PASSWORD", "chroma_secure_2026")
    if credentials.username != api_user or credentials.password != api_password:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials.username

def sanitize_filename(filename: str) -> str:
    """Sanitize filename to prevent path traversal attacks."""
    if not filename:
        raise HTTPException(status_code=400, detail="Filename cannot be empty.")
    
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(
            status_code=400,
            detail="Invalid filename. Directory traversal characters are not allowed."
        )
        
    clean_name = Path(filename).name
    if clean_name != filename:
        raise HTTPException(
            status_code=400,
            detail="Invalid filename. Path traversal detected."
        )
    return clean_name

def sanitize_sample_id(sample_id: str) -> str:
    """Sanitize sample_id to prevent path traversal attacks."""
    if not sample_id:
        raise HTTPException(status_code=400, detail="Sample ID cannot be empty.")
    
    if "/" in sample_id or "\\" in sample_id or ".." in sample_id:
        raise HTTPException(
            status_code=400,
            detail="Invalid Sample ID. Directory traversal characters are not allowed."
        )
        
    clean_id = Path(sample_id).name
    if clean_id != sample_id:
        raise HTTPException(
            status_code=400,
            detail="Invalid Sample ID. Path traversal detected."
        )
    return clean_id


def check_and_convert_proprietary_file(filename: str) -> str:
    """
    If the file is proprietary (e.g. .xms), attempts to find msconvert
    and convert it to .mzML in the RAW_DIR. Returns the new filename.
    """
    import shutil
    import glob
    
    path = RAW_DIR / filename
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"File not found: {filename}")
        
    if path.suffix.lower() in (".raw", ".wiff", ".d"):
        # 1. Search for msconvert
        msconvert_path = shutil.which("msconvert")
        if not msconvert_path:
            # Common paths on Windows
            common_patterns = [
                r"C:\Program Files\ProteoWizard*\msconvert.exe",
                r"C:\Program Files (x86)\ProteoWizard*\msconvert.exe",
                r"C:\Users\*\AppData\Local\ProteoWizard*\msconvert.exe",
                r"C:\Users\*\AppData\Local\Apps\ProteoWizard*\msconvert.exe"
            ]
            for pattern in common_patterns:
                matches = glob.glob(pattern)
                if matches:
                    msconvert_path = matches[0]
                    break
                    
        if not msconvert_path:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Detected proprietary instrument file '{filename}'. To process this format, "
                    "ProteoWizard 'msconvert' is required but was not found on your system. "
                    "Please install ProteoWizard (https://proteowizard.github.io/) and add it to your system PATH."
                )
            )
            
        output_filename = f"{path.stem}.mzML"
        output_path = RAW_DIR / output_filename
        
        # Avoid converting again if already converted and newer than source
        if output_path.exists() and output_path.stat().st_mtime > path.stat().st_mtime:
            print(f"[PREPROCESS] Converted file already exists: {output_filename}")
            return output_filename
            
        print(f"[PREPROCESS] Converting {filename} to {output_filename} using {msconvert_path}...")
        try:
            result = subprocess.run(
                [msconvert_path, str(path), "--mzML", "-o", str(RAW_DIR)],
                capture_output=True, text=True, timeout=300
            )
            if result.returncode != 0:
                raise HTTPException(
                    status_code=500,
                    detail=f"Failed to convert proprietary file {filename} using msconvert: {result.stderr}"
                )
            print(f"[PREPROCESS] Successfully converted {filename} to {output_filename}")
            return output_filename
        except subprocess.TimeoutExpired:
            raise HTTPException(status_code=504, detail=f"Conversion of {filename} using msconvert timed out")
            
    return filename


# ── Serve Dashboard ──────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
def get_dashboard(username: str = Depends(authenticate_user)):
    """Serve the chromatography control center dashboard HTML page."""
    dashboard_path = SCRIPTS_DIR / "dashboard.html"
    if not dashboard_path.exists():
        raise HTTPException(status_code=404, detail="dashboard.html not found in scripts folder")
    return HTMLResponse(content=dashboard_path.read_text(encoding="utf-8"))


# ── Request / Response models ────────────────────────────────────────

class ParseRequest(BaseModel):
    file: str          # filename only, e.g. "sample_01.cdf" — must be in raw_data/
    min_height: float = 100.0
    prominence: float = 50.0

class EnrichRequest(BaseModel):
    sample_id: str
    peaks: list        # list of peak dicts from /run/parse_cdf
    context: str = "GC-MS chromatography, chemical analysis"

class DeconvRequest(BaseModel):
    sample_id: str

class MatchRequest(BaseModel):
    sample_id: str

class AlignRequest(BaseModel):
    sample_ids: list[str]
    reference_sample_id: str = None


# ── Instrument File Upload Endpoint ──────────────────────────────────

@app.post("/upload")
def upload_file(file: UploadFile = File(...), username: str = Depends(authenticate_user)):
    """Upload a raw chromatography file (.cdf or .mzML) to raw_data/ directory."""
    filename = file.filename
    lower_name = filename.lower()
    SUPPORTED_EXTENSIONS = (".cdf", ".mzml", ".xms", ".raw", ".wiff", ".d")
    if not any(lower_name.endswith(ext) for ext in SUPPORTED_EXTENSIONS):
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file format. Supported formats: {', '.join(SUPPORTED_EXTENSIONS)}"
        )

    safe_filename = sanitize_filename(filename)

    # Enforce 300 MB size limit
    MAX_SIZE = 300 * 1024 * 1024
    try:
        file.file.seek(0, 2)
        size = file.file.tell()
        file.file.seek(0)
        if size > MAX_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"File too large. Maximum upload size is 300 MB. Uploaded size: {size / (1024*1024):.2f} MB"
            )
    except Exception as size_err:
        if isinstance(size_err, HTTPException):
            raise size_err
        raise HTTPException(status_code=500, detail=f"Could not check file size: {size_err}")

    output_path = RAW_DIR / safe_filename
    try:
        content = file.file.read()
        output_path.write_bytes(content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to write file: {e}")

    return {
        "status": "ok",
        "filename": safe_filename,
        "message": f"Successfully uploaded {safe_filename} to raw_data/"
    }


# ── Endpoints ────────────────────────────────────────────────────────

@app.get("/health")
def health(username: str = Depends(authenticate_user)):
    """Liveness probe — n8n checks this before starting a workflow."""
    litellm_ok = False
    try:
        r = requests.get(f"{LITELLM_URL}/health/liveness", timeout=10, proxies={"http": None, "https": None})
        litellm_ok = r.status_code == 200
    except Exception as e:
        print(f"[HEALTH CHECK ERROR] Failed to connect to LiteLLM: {e}")
        pass
    return {
        "status": "ok" if litellm_ok else "degraded",
        "server": "pipeline_server",
        "litellm_proxy": "up" if litellm_ok else "down",
        "watcher_active": True,
        "watcher_status": WATCHER_STATUS,
        "raw_data_files": sorted(
            [f.name for f in RAW_DIR.iterdir() if f.is_file() and f.suffix.lower() in (".cdf", ".mzml", ".xms", ".raw", ".wiff", ".d")]
        ),
        "processed_files": sorted([f.name for f in PROC_DIR.glob("*.json")]),
    }


@app.post("/run/parse_cdf")
def run_parse_cdf(req: ParseRequest, username: str = Depends(authenticate_user)):
    """
    Run parse_cdf.py on a raw .cdf file.
    Logic: ALS baseline correction → scipy peak detection → np.trapz integration → JSON.
    Returns: {sample_id, peaks: [{retention_time, peak_area_mAU, baseline_corrected}], ...}
    """
    req.file = sanitize_filename(req.file)
    req.file = check_and_convert_proprietary_file(req.file)
    input_file = RAW_DIR / req.file
    if not req.file.strip():
        raise HTTPException(status_code=400, detail="Filename parameter 'file' is empty.")
    if not input_file.exists() or not input_file.is_file():
        raise HTTPException(status_code=404, detail=f"File not found or is a directory: {req.file}")

    sample_id = input_file.stem
    output_file = PROC_DIR / f"{sample_id}.json"

    try:
        result = subprocess.run(
            [PYTHON, str(SCRIPTS_DIR / "parse_cdf.py"), str(input_file)],
            capture_output=True, text=True, timeout=120, cwd=str(BASE_DIR)
        )
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=504, detail="parse_cdf.py timed out after 120s")

    if result.returncode != 0:
        raise HTTPException(status_code=500, detail={
            "error": "parse_cdf.py failed",
            "stderr": result.stderr[-1000:],
            "stdout": result.stdout[-500:],
        })

    # parse_cdf.py writes JSON to processed/{sample_id}.json and prints summary to stdout
    if not output_file.exists():
        raise HTTPException(status_code=500, detail={
            "error": f"parse_cdf.py ran but did not produce {output_file.name}",
            "stdout": result.stdout[-500:],
            "stderr": result.stderr[-500:],
        })
    try:
        data = json.loads(output_file.read_text())
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail={
            "error": f"{output_file.name} is not valid JSON",
        })

    return {
        "status": "ok",
        "sample_id": sample_id,
        "output_file": str(output_file),
        "peak_count": len(data.get("peaks", [])),
        "data": data,
    }


@app.post("/run/enrich")
def run_enrich(req: EnrichRequest, username: str = Depends(authenticate_user)):
    """
    Call LiteLLM and Claude proxy fallbacks to enrich compound names from peak list.
    Chunks the peak list into batches of 15 to prevent truncation and output limits.
    """
    req.sample_id = sanitize_sample_id(req.sample_id)
    if not req.peaks:
        raise HTTPException(status_code=400, detail="peaks list is empty")

    # Dynamic TMS Context Check
    enrich_context = req.context
    has_tms_matches = any(
        isinstance(p.get("compound_name"), str) and 
        ("tms" in p["compound_name"].lower() or "trimethylsilyl" in p["compound_name"].lower())
        for p in req.peaks
    )
    
    has_tms_spectral_evidence = False
    for p in req.peaks:
        mz_vals = p.get("mz_values", [])
        int_vals = p.get("intensity_values", [])
        if mz_vals and int_vals:
            max_int = max(int_vals) if int_vals else 1.0
            if max_int > 0:
                for mz, val in zip(mz_vals, int_vals):
                    if 72.5 <= mz <= 73.5 and (val / max_int) >= 0.10:
                        has_tms_spectral_evidence = True
                        break
        if has_tms_spectral_evidence:
            break
            
    if (has_tms_matches or has_tms_spectral_evidence) and "trimethylsilyl" not in enrich_context.lower() and "tms" not in enrich_context.lower():
        enrich_context += ", TMS (trimethylsilyl) derivatized sample"

    # Filter only peaks that are actually unidentified or unknown to save API calls and time
    peaks_to_enrich = [p for p in req.peaks if p.get("compound_name") in ("unidentified", "unknown") or not p.get("compound_name")]
    
    # Sort by peak area and height descending to prioritize significant signals
    peaks_to_enrich.sort(
        key=lambda p: (float(p.get("peak_area_mAU") or p.get("area") or 0.0), float(p.get("peak_height_mAU") or p.get("height") or 0.0)),
        reverse=True
    )
    
    # Limit to top N peaks to prevent huge number of API calls and timeouts
    limit_enrich = int(os.environ.get("CHROMA_LIMIT_LLM_PEAKS", "100"))
    peaks_to_enrich = peaks_to_enrich[:limit_enrich]
    
    chunk_size = 15
    enriched_results = []
    chunk_models = []

    # Try to enrich each chunk using the best available fallback model
    for start_i in range(0, len(peaks_to_enrich), chunk_size):
        chunk = peaks_to_enrich[start_i:start_i + chunk_size]
        failed_models = set()
        
        summary_lines = []
        for p in chunk:
            rt_val = p.get('retention_time')
            rt_str = f"{rt_val:.2f}" if isinstance(rt_val, (int, float)) else str(rt_val if rt_val is not None else '?')
            area_val = p.get('peak_area_mAU') or p.get('area')
            area_str = f"{area_val:.1f}" if isinstance(area_val, (int, float)) else str(area_val if area_val is not None else '?')
            name_val = p.get('compound_name', 'unidentified')
            summary_lines.append(f"  Peak {p.get('peak_index')}: RT={rt_str}min, area={area_str} mAU, current_id={name_val}")
        peak_summary = "\n".join(summary_lines)

        prompt = f"""You are a chromatography expert. Given these GC-MS peaks from sample '{req.sample_id}':

{peak_summary}

Context: {enrich_context}

For each peak, suggest:
1. Likely compound class (alcohol, alkane, ester, acid, aromatic, terpene, ketone, amine, unknown)
2. Likely compound name (if current_id is unidentified or unknown, suggest a standard compound name based on RT; otherwise suggest the current_id)
3. Confidence: high/medium/low

Return ONLY a JSON array: [{{"peak_index": 337, "compound_class": "...", "compound_name": "...", "confidence": "..."}}]"""

        chunk_enriched = None
        model_used_for_chunk = None
        
        antigravity_base = os.environ.get("ANTIGRAVITY_BASE_URL", "http://localhost:8080")
        antigravity_model = os.environ.get("ANTIGRAVITY_MODEL", "claude-opus-4-6-thinking")
        
        # Fallback configs to try for this chunk: (type, model_name, timeout)
        # Prioritize fast, high-availability, non-thinking proxy models to avoid timeouts and rate-limit issues
        models_to_try = [
            ("proxy", "gemini-2.5-flash", 20),
            ("proxy", "gemini-2.5-flash-lite", 20),
            ("proxy", "gemini-3-flash-agent", 20),
            ("proxy", "gemini-3.5-flash-low", 20),
            ("proxy", "gemini-2.5-pro", 30),
            ("proxy", antigravity_model, 45),
            ("litellm", "claude-t2", 20),
        ]
        
        for source, model_name, current_timeout in models_to_try:
            model_key = f"{source}:{model_name}"
            if model_key in failed_models:
                continue
            try:
                raw_text = None
                if source == "litellm":
                    r = requests.post(
                        f"{LITELLM_URL}/v1/chat/completions",
                        headers={"Authorization": "Bearer sk-litellm-1234", "Content-Type": "application/json"},
                        json={
                            "model": model_name,
                            "messages": [{"role": "user", "content": prompt}],
                            "max_tokens": 2000,
                            "temperature": 0.2,
                        },
                        timeout=current_timeout,
                        proxies={"http": None, "https": None},
                    )
                    r.raise_for_status()
                    raw_text = r.json()["choices"][0]["message"]["content"].strip()
                else:
                    url = f"{antigravity_base}/messages" if "/v1" in antigravity_base else f"{antigravity_base}/v1/messages"
                    r = requests.post(
                        url,
                        headers={"Content-Type": "application/json"},
                        json={
                            "model": model_name,
                            "messages": [{"role": "user", "content": prompt}],
                            "max_tokens": 2000,
                            "temperature": 0.2,
                        },
                        timeout=current_timeout,
                        proxies={"http": None, "https": None},
                    )
                    r.raise_for_status()
                    response_json = r.json()
                    
                    content = ""
                    if "content" in response_json:
                        content = "".join([c["text"] for c in response_json["content"] if c.get("type") == "text"]).strip()
                    elif "text" in response_json:
                        content = response_json["text"].strip()
                        
                    if not content:
                        thinking_content = "".join([c.get("thinking", "") for c in response_json.get("content", []) if c.get("type") == "thinking"]).strip()
                        if thinking_content:
                            content = thinking_content
                            
                    if content:
                        raw_text = content
                
                if not raw_text:
                    continue

                # Parse the response safely
                json_str = raw_text.strip()
                if "```" in json_str:
                    parts = json_str.split("```")
                    for part in parts:
                        striped_part = part.strip()
                        if striped_part.startswith("[") or striped_part.startswith("{"):
                            json_str = striped_part
                            if json_str.startswith("json"):
                                json_str = json_str[4:].strip()
                            break

                # Slice from first '['/'{' to last ']'/'}'
                start_arr = json_str.find("[")
                end_arr = json_str.rfind("]")
                if start_arr != -1 and end_arr != -1 and end_arr > start_arr:
                    json_str = json_str[start_arr:end_arr+1]
                else:
                    start_obj = json_str.find("{")
                    end_obj = json_str.rfind("}")
                    if start_obj != -1 and end_obj != -1 and end_obj > start_obj:
                        json_str = json_str[start_obj:end_obj+1]

                parsed_val = json.loads(json_str, strict=False)

                if isinstance(parsed_val, dict):
                    for k, v in parsed_val.items():
                        if isinstance(v, list):
                            parsed_val = v
                            break

                if not isinstance(parsed_val, list):
                    parsed_val = [parsed_val]

                # Make sure we got a non-empty list of suggestions
                if len(parsed_val) > 0:
                    chunk_enriched = parsed_val
                    model_used_for_chunk = f"{source} ({model_name})"
                    print(f"[ENRICH] Successfully enriched batch ({start_i+1} to {start_i+len(chunk)}) using {model_used_for_chunk}")
                    break
            except Exception as batch_err:
                print(f"[ENRICH] Model {model_name} failed/parse error for batch ({start_i+1} to {start_i+len(chunk)}): {batch_err}")
                failed_models.add(model_key)


        if not chunk_enriched:
            print(f"[ENRICH ERROR] All fallback models failed/parsed error for batch ({start_i+1} to {start_i+len(chunk)}). Using empty suggestions.")
            chunk_enriched = [{"peak_index": p.get("peak_index"), "compound_class": "parse_error", "compound_name": "unidentified", "confidence": "low"} for p in chunk]
            model_used_for_chunk = "fallback_error"

        enriched_results.extend(chunk_enriched)
        chunk_models.append(model_used_for_chunk)

    model_used = ", ".join(sorted(list(set(chunk_models)))) if chunk_models else "none (all peaks matched)"

    # Merge enrichment back into peaks matching the actual peak_index
    enriched_map = {e.get("peak_index"): e for e in enriched_results if e.get("peak_index") is not None}
    peaks_enriched = []
    for p in req.peaks:
        pidx = p.get("peak_index")
        enrich_data = enriched_map.get(pidx, {})
        
        # Merge values directly into the peak copy
        merged = {**p}
        
        # Suggest class if not already matched or if unknown
        if enrich_data.get("compound_class") and (p.get("compound_class") == "unknown" or not p.get("compound_class")):
            merged["compound_class"] = enrich_data["compound_class"]
            
        # Suggest compound name if unidentified
        if enrich_data.get("compound_name") and (p.get("compound_name") == "unidentified" or not p.get("compound_name")):
            merged["compound_name"] = enrich_data["compound_name"]
            
        # Record enrichment confidence
        if enrich_data.get("confidence"):
            merged["enrich_confidence"] = enrich_data["confidence"]
            
        peaks_enriched.append(merged)

    # Save enriched result
    output_file = PROC_DIR / f"{req.sample_id}_enriched.json"
    output_file.write_text(json.dumps({
        "sample_id": req.sample_id,
        "enriched_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "model_used": model_used,
        "peaks": peaks_enriched,
    }, indent=2))

    return {
        "status": "ok",
        "sample_id": req.sample_id,
        "peaks_enriched": len(peaks_enriched),
        "output_file": str(output_file),
        "peaks": peaks_enriched,
    }



@app.post("/run/deconv")
def run_deconv(req: DeconvRequest, username: str = Depends(authenticate_user)):
    """
    Run gnn_deconv.py on a processed sample peaks JSON.
    Uses Graph Neural Network GCN to deconvolve mixed signals.
    """
    req.sample_id = sanitize_sample_id(req.sample_id)
    input_file = PROC_DIR / f"{req.sample_id}.json"
    if not input_file.exists():
        raise HTTPException(status_code=404, detail=f"Processed peak file not found: {req.sample_id}.json")

    output_file = PROC_DIR / f"{req.sample_id}_deconvolved.json"

    try:
        result = subprocess.run(
            [PYTHON, str(SCRIPTS_DIR / "gnn_deconv.py"), str(input_file)],
            capture_output=True, text=True, timeout=120, cwd=str(BASE_DIR)
        )
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=504, detail="gnn_deconv.py timed out after 120s")

    if result.returncode != 0:
        raise HTTPException(status_code=500, detail={
            "error": "gnn_deconv.py failed",
            "stderr": result.stderr[-1000:],
            "stdout": result.stdout[-500:],
        })

    if not output_file.exists():
        raise HTTPException(status_code=500, detail={
            "error": f"gnn_deconv.py ran but did not produce {output_file.name}",
            "stdout": result.stdout[-500:],
            "stderr": result.stderr[-500:],
        })

    try:
        data = json.loads(output_file.read_text())
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail={
            "error": f"{output_file.name} is not valid JSON",
        })

    return data


@app.post("/run/match")
def run_match(req: MatchRequest, username: str = Depends(authenticate_user)):
    """
    Run spectral_match.py on deconvolved or raw peaks JSON.
    """
    req.sample_id = sanitize_sample_id(req.sample_id)
    # Prefer deconvolved peaks if they exist
    input_file = PROC_DIR / f"{req.sample_id}_deconvolved.json"
    if not input_file.exists():
        input_file = PROC_DIR / f"{req.sample_id}.json"

    if not input_file.exists():
        raise HTTPException(status_code=404, detail=f"No peak JSON file found to match for sample: {req.sample_id}")

    output_file = PROC_DIR / f"{req.sample_id}_matched.json"

    try:
        result = subprocess.run(
            [PYTHON, str(SCRIPTS_DIR / "spectral_match.py"), str(input_file)],
            capture_output=True, text=True, timeout=300, cwd=str(BASE_DIR)
        )
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=504, detail="spectral_match.py timed out after 300s")

    if result.returncode != 0:
        raise HTTPException(status_code=500, detail={
            "error": "spectral_match.py failed",
            "stderr": result.stderr[-1000:],
            "stdout": result.stdout[-500:],
        })

    if not output_file.exists():
        raise HTTPException(status_code=500, detail={
            "error": f"spectral_match.py ran but did not produce {output_file.name}",
            "stdout": result.stdout[-500:],
            "stderr": result.stderr[-500:],
        })

    try:
        data = json.loads(output_file.read_text())
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail={
            "error": f"{output_file.name} is not valid JSON",
        })

    # Ensure peaks is returned if not written to matched.json yet
    if "peaks" not in data:
        try:
            data["peaks"] = json.loads(input_file.read_text()).get("peaks", [])
        except Exception:
            pass

    return data


@app.get("/results/{sample_id}")
def get_results(sample_id: str, username: str = Depends(authenticate_user)):
    """Fetch a previously processed result from processed/ folder.
    Combines the most advanced data (enriched > matched > deconvolved > parsed)
    and ensures all metadata (deconvolution status, spectral matches) are merged."""
    sample_id = sanitize_sample_id(sample_id)
    base_id = sample_id
    for suffix in ["_enriched", "_matched", "_deconvolved"]:
        if base_id.endswith(suffix):
            base_id = base_id[:-len(suffix)]
            break

    # Determine files
    parsed_file = PROC_DIR / f"{base_id}.json"
    deconv_file = PROC_DIR / f"{base_id}_deconvolved.json"
    matched_file = PROC_DIR / f"{base_id}_matched.json"
    enriched_file = PROC_DIR / f"{base_id}_enriched.json"

    # Base payload structure
    data = {}
    
    # 1. Start with the most basic parsed peak coordinate info
    if parsed_file.exists():
        try:
            data = json.loads(parsed_file.read_text())
        except Exception:
            pass
            
    # If not even parsed exists, try to load deconv or matched as baseline
    if not data:
        for f in [deconv_file, matched_file, enriched_file]:
            if f.exists():
                try:
                    data = json.loads(f.read_text())
                    break
                except Exception:
                    pass

    if not data:
        raise HTTPException(status_code=404, detail=f"No results found for sample: {sample_id}")

    # Ensure deconvolution and matches keys exist
    if "deconvolution" not in data:
        data["deconvolution"] = "incomplete"

    # 2. Layer deconvolution
    if deconv_file.exists():
        try:
            deconv_data = json.loads(deconv_file.read_text())
            data["deconvolution"] = deconv_data.get("deconvolution", "complete")
            data["n_regions"] = deconv_data.get("n_regions")
            data["n_coeluting_regions"] = deconv_data.get("n_coeluting_regions")
            if "peaks" in deconv_data:
                data["peaks"] = deconv_data["peaks"]
        except Exception:
            pass

    # 3. Layer matching
    if matched_file.exists():
        try:
            matched_data = json.loads(matched_file.read_text())
            data["matches"] = matched_data.get("matches", [])
            data["match_count"] = matched_data.get("match_count", len(data.get("matches", [])))
            data["total_peaks"] = matched_data.get("total_peaks", len(data.get("peaks", [])))
            if "peaks" in matched_data:
                data["peaks"] = matched_data["peaks"]
            
            # Merge matches list into peaks if not already present (for backwards compatibility with historical cache files)
            matches_map = {m.get("peak_index"): m for m in data.get("matches", []) if m.get("peak_index") is not None}
            for p in data.get("peaks", []):
                p_idx = p.get("peak_index")
                if p_idx in matches_map:
                    match_info = matches_map[p_idx]
                    p["compound_name"] = match_info.get("compound_name")
                    p["compound_class"] = match_info.get("compound_class")
                    p["cosine_similarity"] = match_info.get("cosine_similarity")
                    p["n_fragments_matched"] = match_info.get("n_fragments_matched")
                else:
                    if "compound_name" not in p:
                        p["compound_name"] = "unidentified"
                    if "compound_class" not in p:
                        p["compound_class"] = "unknown"
                    if "cosine_similarity" not in p:
                        p["cosine_similarity"] = "N/A"
        except Exception:
            pass

    # 4. Layer enrichment
    if enriched_file.exists():
        try:
            enriched_data = json.loads(enriched_file.read_text())
            data["enriched_at"] = enriched_data.get("enriched_at")
            data["model_used"] = enriched_data.get("model_used")
            if "peaks" in enriched_data:
                # Merge enriched peaks data
                enriched_peaks_map = {p.get("peak_index"): p for p in enriched_data["peaks"]}
                for p in data.get("peaks", []):
                    p_idx = p.get("peak_index")
                    if p_idx in enriched_peaks_map:
                        p["compound_class"] = enriched_peaks_map[p_idx].get("compound_class")
                        p["confidence"] = enriched_peaks_map[p_idx].get("confidence")
                        if enriched_peaks_map[p_idx].get("compound_name"):
                            p["compound_name"] = enriched_peaks_map[p_idx]["compound_name"]
        except Exception:
            pass

    return data


def load_processed_data(sample_id: str) -> dict:
    """Helper to load the most advanced peak data and retention times for a sample."""
    base_id = sample_id
    for suffix in ["_enriched", "_matched", "_deconvolved"]:
        if base_id.endswith(suffix):
            base_id = base_id[:-len(suffix)]
            break

    parsed_file = PROC_DIR / f"{base_id}.json"
    deconv_file = PROC_DIR / f"{base_id}_deconvolved.json"
    matched_file = PROC_DIR / f"{base_id}_matched.json"
    enriched_file = PROC_DIR / f"{base_id}_enriched.json"

    data = {}
    if parsed_file.exists():
        try:
            data = json.loads(parsed_file.read_text())
        except Exception:
            pass
            
    if not data:
        for f in [deconv_file, matched_file, enriched_file]:
            if f.exists():
                try:
                    data = json.loads(f.read_text())
                    break
                except Exception:
                    pass

    if not data:
        raise HTTPException(status_code=404, detail=f"No results found for sample: {sample_id}")

    # Layer peaks info sequentially (parsed < deconvolved < matched < enriched)
    if deconv_file.exists():
        try:
            deconv_data = json.loads(deconv_file.read_text())
            if "peaks" in deconv_data:
                data["peaks"] = deconv_data["peaks"]
        except Exception:
            pass

    if matched_file.exists():
        try:
            matched_data = json.loads(matched_file.read_text())
            if "peaks" in matched_data:
                data["peaks"] = matched_data["peaks"]
        except Exception:
            pass

    if enriched_file.exists():
        try:
            enriched_data = json.loads(enriched_file.read_text())
            if "peaks" in enriched_data:
                # Merge enriched peaks data
                enriched_peaks_map = {p.get("peak_index"): p for p in enriched_data["peaks"]}
                for p in data.get("peaks", []):
                    p_idx = p.get("peak_index")
                    if p_idx in enriched_peaks_map:
                        p["compound_class"] = enriched_peaks_map[p_idx].get("compound_class")
                        p["confidence"] = enriched_peaks_map[p_idx].get("confidence")
                        if enriched_peaks_map[p_idx].get("compound_name"):
                            p["compound_name"] = enriched_peaks_map[p_idx]["compound_name"]
        except Exception:
            pass

    return data


def _perform_alignment(req: AlignRequest) -> dict:
    """Helper function to perform alignment calculations for clean reuse."""
    import numpy as np
    import sys
    
    # Make sure SCRIPTS_DIR is in path
    if str(SCRIPTS_DIR) not in sys.path:
        sys.path.append(str(SCRIPTS_DIR))
    import alignment

    if not req.sample_ids or len(req.sample_ids) < 2:
        raise HTTPException(status_code=400, detail="At least two sample IDs must be provided for alignment.")

    # Sanitize and load all sample data
    samples_data = {}
    for sid in req.sample_ids:
        safe_sid = sanitize_sample_id(sid)
        try:
            samples_data[sid] = load_processed_data(safe_sid)
        except HTTPException as e:
            raise e
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to load data for sample {sid}: {e}")

    # Determine reference sample ID
    ref_sid = req.reference_sample_id
    if ref_sid:
        if ref_sid not in samples_data:
            safe_ref_sid = sanitize_sample_id(ref_sid)
            if safe_ref_sid in samples_data:
                ref_sid = safe_ref_sid
            else:
                raise HTTPException(status_code=400, detail=f"Reference sample ID '{ref_sid}' not found in sample_ids.")
    else:
        # Choose run with the largest total peak area
        max_area = -1.0
        best_sid = None
        for sid, sdata in samples_data.items():
            peaks = sdata.get("peaks", [])
            total_area = sum(float(p.get("peak_area_mAU", 0.0)) for p in peaks)
            if total_area > max_area:
                max_area = total_area
                best_sid = sid
        ref_sid = best_sid

    ref_data = samples_data[ref_sid]
    rt_ref_orig = np.array(ref_data["retention_time"], dtype=np.float64)
    peaks_ref = ref_data.get("peaks", [])

    # Downsample reference grid to 1000 points for CPU-friendly O(N) alignment
    rt_ref = np.linspace(rt_ref_orig.min(), rt_ref_orig.max(), 1000)

    # Reconstruct reference chromatogram signal
    try:
        int_ref = alignment.reconstruct_chromatogram_from_peaks(rt_ref, peaks_ref)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to reconstruct chromatogram for reference {ref_sid}: {e}")

    aligned_runs = {}

    for sid, sdata in samples_data.items():
        if sid == ref_sid:
            aligned_runs[sid] = {
                "aligned_intensity": [float(val) for val in int_ref],
                "peaks": peaks_ref,
                "warping_path": [[i, i] for i in range(len(rt_ref))]
            }
            continue

        rt_target_orig = np.array(sdata["retention_time"], dtype=np.float64)
        peaks_target = sdata.get("peaks", [])

        # Downsample target grid to 1000 points
        rt_target = np.linspace(rt_target_orig.min(), rt_target_orig.max(), 1000)

        # Reconstruct target chromatogram signal
        try:
            int_target = alignment.reconstruct_chromatogram_from_peaks(rt_target, peaks_target)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to reconstruct chromatogram for target {sid}: {e}")

        # Align signals
        try:
            aligned_int, path = alignment.align_chromatogram_signals(
                rt_ref, int_ref, rt_target, int_target, window_seconds=15.0
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Signal alignment failed for {sid}: {e}")

        # Map peaks
        try:
            _, updated_target = alignment.map_and_align_peaks(
                peaks_ref, peaks_target, rt_ref, rt_target, path, rt_tolerance_min=0.05
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Peak alignment failed for {sid}: {e}")

        aligned_runs[sid] = {
            "aligned_intensity": [float(val) for val in aligned_int],
            "peaks": updated_target,
            "warping_path": [[int(i), int(j)] for i, j in path]
        }

    return {
        "status": "ok",
        "reference_sample_id": ref_sid,
        "reference_rt": [float(t) for t in rt_ref],
        "reference_intensity": [float(val) for val in int_ref],
        "aligned_runs": aligned_runs
    }


@app.post("/run/align")
def run_align(req: AlignRequest, username: str = Depends(authenticate_user)):
    """Perform multi-sample chromatographic alignment and retention time drift correction."""
    return _perform_alignment(req)


@app.post("/export/comparison")
def export_comparison(req: AlignRequest, username: str = Depends(authenticate_user)):
    """
    Generate and download a side-by-side comparative Excel report
    for multiple aligned runs.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    align_res = _perform_alignment(req)
    ref_sid = align_res["reference_sample_id"]
    aligned_runs = align_res["aligned_runs"]
    
    ref_peaks = aligned_runs[ref_sid]["peaks"]
    rows = []
    
    ref_peak_map = {}
    for rp in ref_peaks:
        row = {
            "compound_name": rp.get("compound_name") or "unidentified",
            "compound_class": rp.get("compound_class") or "unknown",
            "ref_rt": rp.get("retention_time"),
            "ref_peak_index": rp.get("peak_index"),
            "peaks_by_sample": {ref_sid: rp}
        }
        rows.append(row)
        ref_peak_map[rp["peak_index"]] = row

    unmatched_target_rows = []
    for sid, run_data in aligned_runs.items():
        if sid == ref_sid:
            continue
        
        target_peaks = run_data["peaks"]
        for tp in target_peaks:
            ref_idx = tp.get("matched_peak_index")
            if ref_idx is not None and ref_idx in ref_peak_map:
                ref_peak_map[ref_idx]["peaks_by_sample"][sid] = tp
            else:
                unmatched_row = {
                    "compound_name": tp.get("compound_name") or "unidentified",
                    "compound_class": tp.get("compound_class") or "unknown",
                    "ref_rt": tp.get("aligned_rt"),
                    "ref_peak_index": None,
                    "peaks_by_sample": {sid: tp}
                }
                unmatched_target_rows.append(unmatched_row)

    unmatched_target_rows.sort(key=lambda x: x["ref_rt"] if x["ref_rt"] is not None else 9999.0)
    all_rows = rows + unmatched_target_rows

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample Comparison"

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    
    data_font = Font(name="Calibri", size=10)
    bold_data_font = Font(name="Calibri", bold=True, size=10)
    data_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    ref_run_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    sample_ids = req.sample_ids
    ordered_samples = [ref_sid] + [sid for sid in sample_ids if sid != ref_sid]
    num_samples = len(ordered_samples)
    total_cols = 4 + (num_samples * 3)

    title_range = f"A1:{get_column_letter(total_cols)}1"
    ws.merge_cells(title_range)
    title_cell = ws["A1"]
    title_cell.value = "CHROMA-AGENT-ALPHA — Multi-Sample Chromatographic Alignment & Comparison"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    meta_range = f"A2:{get_column_letter(total_cols)}2"
    ws.merge_cells(meta_range)
    meta_cell = ws["A2"]
    meta_cell.value = f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S')} | Alignment Window: ±15s | Reference: {ref_sid} | Total Runs: {num_samples}"
    meta_cell.font = Font(name="Calibri", size=9, italic=True, color="808080")
    meta_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[4].height = 24
    ws.merge_cells("A4:D4")
    c_hdr = ws.cell(row=4, column=1, value="Peak Identification & Timeline")
    c_hdr.font = header_font
    c_hdr.fill = header_fill
    c_hdr.alignment = header_align
    c_hdr.border = thin_border
    
    for col in range(2, 5):
        ws.cell(row=4, column=col).border = thin_border
        ws.cell(row=4, column=col).fill = header_fill

    current_col = 5
    for sid in ordered_samples:
        col_letter_start = get_column_letter(current_col)
        col_letter_end = get_column_letter(current_col + 2)
        merge_range = f"{col_letter_start}4:{col_letter_end}4"
        ws.merge_cells(merge_range)
        
        lbl = f"Run: {sid}"
        if sid == ref_sid:
            lbl += " [REF]"
            
        s_hdr = ws.cell(row=4, column=current_col, value=lbl)
        s_hdr.font = header_font
        s_hdr.fill = subheader_fill if sid != ref_sid else header_fill
        s_hdr.alignment = header_align
        s_hdr.border = thin_border
        
        for col in range(current_col + 1, current_col + 3):
            c = ws.cell(row=4, column=col)
            c.border = thin_border
            c.fill = subheader_fill if sid != ref_sid else header_fill
            
        current_col += 3

    ws.row_dimensions[5].height = 20
    sub_headers = ["Compound Name", "Class", "Ref Peak Index", "Ref RT (min)"]
    for sid in ordered_samples:
        sub_headers.extend(["Area (mAU·min)", "Height (mAU)", "RT (min)"])
        
    for col_idx, sh in enumerate(sub_headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=sh)
        cell.font = header_font
        is_ref_metric = (5 <= col_idx <= 7) or (col_idx <= 4)
        cell.fill = header_fill if is_ref_metric else subheader_fill
        cell.alignment = header_align
        cell.border = thin_border

    current_row = 6
    for row_data in all_rows:
        ws.row_dimensions[current_row].height = 18
        
        c_name = ws.cell(row=current_row, column=1, value=row_data["compound_name"])
        c_name.font = bold_data_font if row_data["ref_peak_index"] is not None else data_font
        c_name.alignment = left_align
        c_name.border = thin_border
        
        c_class = ws.cell(row=current_row, column=2, value=row_data["compound_class"])
        c_class.font = data_font
        c_class.alignment = left_align
        c_class.border = thin_border
        
        ref_pi_val = row_data["ref_peak_index"]
        c_rpi = ws.cell(row=current_row, column=3, value=ref_pi_val if ref_pi_val is not None else "N/A")
        c_rpi.font = data_font
        c_rpi.alignment = data_align
        c_rpi.border = thin_border
        
        ref_rt_val = row_data["ref_rt"]
        c_rrt = ws.cell(row=current_row, column=4, value=round(ref_rt_val, 4) if ref_rt_val is not None else "N/A")
        c_rrt.font = data_font
        c_rrt.alignment = data_align
        c_rrt.border = thin_border
        
        col_offset = 5
        for sid in ordered_samples:
            peak = row_data["peaks_by_sample"].get(sid)
            is_ref = (sid == ref_sid)
            
            if peak is not None:
                area_cell = ws.cell(row=current_row, column=col_offset, value=float(peak.get("peak_area_mAU", 0.0)))
                area_cell.number_format = "0.00"
                
                height_cell = ws.cell(row=current_row, column=col_offset + 1, value=float(peak.get("peak_height_mAU", 0.0)))
                height_cell.number_format = "0.0"
                
                rt_val = peak.get("retention_time")
                rt_cell = ws.cell(row=current_row, column=col_offset + 2, value=float(rt_val))
                rt_cell.number_format = "0.000"
            else:
                area_cell = ws.cell(row=current_row, column=col_offset, value=0.0)
                area_cell.number_format = "0.00"
                
                height_cell = ws.cell(row=current_row, column=col_offset + 1, value=0.0)
                height_cell.number_format = "0.0"
                
                rt_cell = ws.cell(row=current_row, column=col_offset + 2, value="-")
            
            for cell in [area_cell, height_cell, rt_cell]:
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                if is_ref:
                    cell.fill = ref_run_fill
                    
            col_offset += 3
            
        current_row += 1

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 3 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16

    output_path = RESULTS_DIR / "comparison_results.xlsx"
    wb.save(output_path)
    
    return FileResponse(
        path=str(output_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="comparison_results.xlsx"
    )



def create_excel_report(
    sample_id: str,
    sample_weight: float = 0.1000,
    api_conc: float = 1.0000,
    calibrations: dict = None
) -> Path:
    """Generate a formatted Excel (.xlsx) report for a processed sample with GC quantification.
    Merges peaks + match data into a formatted spreadsheet saved to RESULTS_DIR."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Strip suffixes to find the base ID
    base_id = sample_id
    for suffix in ["_enriched", "_matched", "_deconvolved"]:
        if base_id.endswith(suffix):
            base_id = base_id[:-len(suffix)]
            break

    # --- Load the best available data ---
    enriched_file = PROC_DIR / f"{base_id}_enriched.json"
    matched_file = PROC_DIR / f"{base_id}_matched.json"
    deconv_file = PROC_DIR / f"{base_id}_deconvolved.json"
    base_file = PROC_DIR / f"{base_id}.json"

    peaks = []
    matches = []
    source_label = "raw"

    if enriched_file.exists():
        data = json.loads(enriched_file.read_text())
        peaks = data.get("peaks", [])
        source_label = "enriched"
        # Load matches from matched file if exists
        if matched_file.exists():
            try:
                matches = json.loads(matched_file.read_text()).get("matches", [])
            except Exception:
                pass
    elif matched_file.exists():
        data = json.loads(matched_file.read_text())
        peaks = data.get("peaks", [])
        matches = data.get("matches", [])
        source_label = "matched"
    elif deconv_file.exists():
        data = json.loads(deconv_file.read_text())
        peaks = data.get("peaks", [])
        source_label = "deconvolved"
    elif base_file.exists():
        data = json.loads(base_file.read_text())
        peaks = data.get("peaks", [])
        source_label = "parsed"
    else:
        raise HTTPException(status_code=404, detail=f"No processed data found for sample: {base_id}")

    if not peaks:
        raise HTTPException(status_code=404, detail=f"No peaks found in data for sample: {base_id}")

    # Build match lookup
    match_map = {}
    for m in matches:
        match_map[m.get("peak_index")] = m

    # --- Create workbook ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Peak Results"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(horizontal="center", vertical="center")
    match_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    nomatch_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    ai_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Alert styles for compliance failures (USP <467> Out of Specification)
    fail_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid") # soft red/orange alert
    fail_font = Font(name="Calibri", bold=True, color="C00000", size=10) # bold dark red text
    pass_font = Font(name="Calibri", bold=True, color="375623", size=10) # bold dark green text

    # Title row (merged T columns wide now)
    ws.merge_cells("A1:T1")
    title_cell = ws["A1"]
    title_cell.value = f"CHROMA-AGENT-ALPHA — Peak Quantification & Regulatory Compliance Report: {base_id}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Metadata row
    ws.merge_cells("A2:T2")
    meta_cell = ws["A2"]
    meta_cell.value = f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S')} | Sample Weight: {sample_weight:.4f} g | API Conc: {api_conc:.4f} mg/mL | Total Peaks: {len(peaks)}"
    meta_cell.font = Font(name="Calibri", size=9, italic=True, color="808080")
    meta_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "Peak Index", "RT (min)", "Height (mAU)", "Area (mAU·min)", "GNN Purity", 
        "Corrected Area (mAU·min)", "Resolution", "Tailing Factor", "Theoretical Plates", "Signal to Noise (S/N)",
        "Compound ID", "Compound Class", "Match Score", "Fragments Matched",
        "Std Area (rStd)", "Response Factor", "USP Limit (ppm)", "Quantified ppm", "Solvent Content (%)", "Compliance"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    USP_467_LIMITS = {
        "benzene": 2.0,
        "carbon tetrachloride": 4.0,
        "1,2-dichloroethane": 5.0,
        "1,1-dichloroethene": 8.0,
        "1,1,1-trichloroethane": 10.0,
        "methanol": 3000.0,
        "acetonitrile": 410.0,
        "dichloromethane": 600.0,
        "hexane": 290.0,
        "toluene": 890.0,
        "ethanol": 5000.0,
        "acetone": 5000.0,
        "ethyl acetate": 5000.0,
        "tetrahydrofuran": 720.0,
        "isopropanol": 5000.0,
        "propanol": 5000.0,
        "1-propanol": 5000.0,
        "2-propanol": 5000.0,
        "butanol": 5000.0,
        "1-butanol": 5000.0,
        "2-butanol": 5000.0,
        "isobutanol": 5000.0
    }

    # Data rows
    for row_idx, p in enumerate(peaks, 5):
        pidx = p.get("peak_index", "")
        match = match_map.get(pidx)
        
        has_ai = (not match) and p.get("compound_name") and p.get("compound_name") != "unidentified"
        row_fill = match_fill if match else (ai_fill if has_ai else nomatch_fill)

        comp_name = match["compound_name"] if match else (p.get("compound_name") if p.get("compound_name") else "unidentified")
        comp_class = match["compound_class"] if match else (p.get("compound_class") if p.get("compound_class") else "unknown")
        
        if match:
            score = round(match["cosine_similarity"], 4)
            frags = match.get("n_fragments_matched", "N/A")
        elif has_ai:
            score = f"AI: {p.get('confidence', p.get('enrich_confidence', 'low'))}"
            frags = "AI"
        else:
            score = "N/A"
            frags = "N/A"

        # GNN Purity formatting (purity is 1.0 for pure peaks, component_purity for coeluting, fallback to 1.0 if not deconvolved)
        purity_val = p.get("component_purity")
        if purity_val is None:
            purity_val = 1.0 if not p.get("coeluting", False) else 0.5
        purity_display = round(purity_val, 3)

        # Standard calibration calculations
        r_std = "N/A"
        rf_val = "N/A"
        usp_limit = "N/A"
        ppm_val = "N/A"
        content_pct = "N/A"
        compliance = "N/A"
        
        # Resolve target name
        norm_name = comp_name.lower().strip()
        
        # Look up USP limit
        for k, v in USP_467_LIMITS.items():
            if k in norm_name:
                usp_limit = v
                break
                
        # If calibrations are supplied and this compound matches
        if calibrations:
            matched_calib_key = None
            for calib_key in calibrations.keys():
                if calib_key.lower().strip() in norm_name:
                    matched_calib_key = calib_key
                    break
            
            if matched_calib_key:
                cal = calibrations[matched_calib_key]
                c_std = float(cal.get("CStd", 0))
                r_std_val = float(cal.get("rStd", 0))
                
                if c_std > 0 and r_std_val > 0:
                    r_std = r_std_val
                    rf_val = round(r_std_val / c_std, 4)
                    
                    # ppm = rSpl / (RF * Wspl) - Corrected using GNN Purity to avoid co-elution errors
                    r_spl = p.get("peak_area_mAU", 0.0) * purity_val
                    if rf_val > 0 and sample_weight > 0:
                        ppm_computed = r_spl / (rf_val * sample_weight)
                        ppm_val = round(ppm_computed, 2)
                        
                        # Content % = (rSpl * CStd * 0.1) / (Csample * rStd)
                        if api_conc > 0:
                            pct_computed = (r_spl * c_std * 0.1) / (api_conc * r_std_val)
                            content_pct = round(pct_computed, 4)
                            
                        # Check compliance
                        if isinstance(usp_limit, (int, float)):
                            if ppm_computed > usp_limit:
                                compliance = "FAIL"
                                row_fill = fail_fill # Mark row in fail color
                            else:
                                compliance = "PASS"
                        else:
                            compliance = "PASS"

        values = [
            pidx,
            round(p.get("retention_time", 0), 4),
            round(p.get("peak_height_mAU", 0), 2),
            round(p.get("peak_area_mAU", 0), 2),
            purity_display,
            round(p.get("peak_area_mAU", 0.0) * purity_val, 2), # Corrected Area
            round(p.get("resolution", 0.0), 2),
            round(p.get("tailing_factor", 1.0), 3),
            int(p.get("theoretical_plates", 0)),
            round(p.get("signal_to_noise", 0.0), 1),
            comp_name,
            comp_class,
            score,
            frags,
            r_std,
            rf_val,
            usp_limit,
            ppm_val,
            content_pct,
            compliance
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            cell.fill = row_fill
            
            # Format compliance text font specially
            if col_idx == 20: # Compliance column
                if val == "FAIL":
                    cell.font = fail_font
                elif val == "PASS":
                    cell.font = pass_font

    # Column widths
    col_widths = [12, 12, 14, 16, 12, 24, 12, 14, 18, 20, 35, 18, 18, 18, 20, 16, 16, 16, 18, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = "A5"

    # Count statistics for Summary sheet
    spectral_count = 0
    ai_count = 0
    unidentified_count = 0
    passed_count = 0
    failed_count = 0
    
    for p in peaks:
        pidx = p.get("peak_index")
        if pidx in match_map:
            spectral_count += 1
        elif p.get("compound_name") and p.get("compound_name") != "unidentified":
            ai_count += 1
        else:
            unidentified_count += 1
            
        # Recalculate compliance for stats
        comp_name_str = match_map[pidx]["compound_name"] if pidx in match_map else (p.get("compound_name") if p.get("compound_name") else "unidentified")
        norm_name_str = comp_name_str.lower().strip()
        
        limit_val = "N/A"
        for k, v in USP_467_LIMITS.items():
            if k in norm_name_str:
                limit_val = v
                break
                
        if calibrations and isinstance(limit_val, (int, float)):
            matched_calib_key = None
            for calib_key in calibrations.keys():
                if calib_key.lower().strip() in norm_name_str:
                    matched_calib_key = calib_key
                    break
            
            if matched_calib_key:
                cal = calibrations[matched_calib_key]
                c_std = float(cal.get("CStd", 0))
                r_std_val = float(cal.get("rStd", 0))
                if c_std > 0 and r_std_val > 0:
                    rf = r_std_val / c_std
                    
                    # Resolve GNN purity for summary stats compliance consistency
                    purity_val_c = p.get("component_purity")
                    if purity_val_c is None:
                        purity_val_c = 1.0 if not p.get("coeluting", False) else 0.5
                        
                    r_spl = p.get("peak_area_mAU", 0.0) * purity_val_c
                    if rf > 0 and sample_weight > 0:
                        ppm_c = r_spl / (rf * sample_weight)
                        if ppm_c > limit_val:
                            failed_count += 1
                        else:
                            passed_count += 1

    # Summary sheet
    ws2 = wb.create_sheet(title="Summary")
    ws2["A1"] = "CHROMA-AGENT-ALPHA Pipeline Summary"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    
    summary_data = [
        ("Sample ID", base_id),
        ("Total Peaks", len(peaks)),
        ("Spectral Matches", spectral_count),
        ("AI Fallback Identifications", ai_count),
        ("Unidentified Peaks", unidentified_count),
        ("Spectral Match Rate", f"{spectral_count/len(peaks)*100:.1f}%" if peaks else "0%"),
        ("Overall Identification Rate", f"{(spectral_count + ai_count)/len(peaks)*100:.1f}%" if peaks else "0%"),
        ("USP <467> Passed Peaks", passed_count),
        ("USP <467> Failed Peaks (OOS)", failed_count),
        ("Quantification Weight (WSpl)", f"{sample_weight:.4f} g"),
        ("API Concentration (CSample)", f"{api_conc:.4f} mg/mL"),
        ("Pipeline Stage", source_label),
        ("Export Time", time.strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for r_idx, (label, value) in enumerate(summary_data, 3):
        ws2.cell(row=r_idx, column=1, value=label).font = Font(name="Calibri", bold=True, size=10)
        ws2.cell(row=r_idx, column=2, value=value).font = Font(name="Calibri", size=10)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 30

    output_path = RESULTS_DIR / f"{base_id}_results.xlsx"
    wb.save(str(output_path))
    return output_path


@app.get("/export/{sample_id}")
def export_excel(
    sample_id: str,
    weight: float = 0.1000,
    api_conc: float = 1.0000,
    calibrations: str = None,
    username: str = Depends(authenticate_user)
):
    """Generate and download an Excel (.xlsx) report for a processed sample.
    Supports optional weight, API concentration, and calibrations JSON string."""
    sample_id = sanitize_sample_id(sample_id)
    try:
        calib_dict = {}
        if calibrations:
            try:
                calib_dict = json.loads(calibrations)
            except Exception as e:
                print(f"[EXPORT] Warning: Failed to parse calibrations JSON: {e}")
                
        output_path = create_excel_report(
            sample_id, 
            sample_weight=weight, 
            api_conc=api_conc, 
            calibrations=calib_dict
        )
        
        # Update Zarr store folder with the user's customized calibration and weight parameters for FAIR alignment
        try:
            import sys
            sys.path.append(str(SCRIPTS_DIR))
            import data_store
            
            # Identify raw instrument file
            base_id = sample_id
            for suffix in ["_enriched", "_matched", "_deconvolved"]:
                if base_id.endswith(suffix):
                    base_id = base_id[:-len(suffix)]
                    break
                    
            input_file = RAW_DIR / f"{base_id}.cdf"
            if not input_file.exists():
                input_file = RAW_DIR / f"{base_id}.mzML"
            if not input_file.exists():
                input_file = RAW_DIR / f"{base_id}.CDF"
                
            if input_file.exists():
                zarr_path = data_store.save_run_to_zarr(
                    base_id,
                    str(input_file),
                    sample_weight=weight,
                    api_conc=api_conc,
                    calibrations=calib_dict
                )
                print(f"[EXPORT ZARR UPDATE] Successfully updated Zarr store: {zarr_path}")
        except Exception as zarr_err:
            print(f"[EXPORT ZARR UPDATE ERROR] Could not update Zarr archive: {zarr_err}")

        return FileResponse(
            path=str(output_path),
            filename=output_path.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/run/full_pipeline")
def run_full_pipeline(req: ParseRequest, username: str = Depends(authenticate_user)):
    """Run the full pipeline: parse_cdf → gnn_deconv → spectral_match in sequence."""
    req.file = sanitize_filename(req.file)
    req.file = check_and_convert_proprietary_file(req.file)
    # Stage 1: Parse CDF
    input_file = RAW_DIR / req.file
    if not req.file.strip():
        raise HTTPException(status_code=400, detail="Filename parameter 'file' is empty.")
    if not input_file.exists() or not input_file.is_file():
        raise HTTPException(status_code=404, detail=f"File not found: {req.file}")

    sample_id = input_file.stem
    stages_completed = []
    
    # Initialize Watcher Status for the manual run so the progress bar updates live
    WATCHER_STATUS["active"] = True
    WATCHER_STATUS["current_file"] = req.file
    WATCHER_STATUS["stage"] = "parsing"
    WATCHER_STATUS["stages_completed"] = []
    WATCHER_STATUS["last_error"] = None

    try:
        # Parse
        try:
            result = subprocess.run(
                [PYTHON, str(SCRIPTS_DIR / "parse_cdf.py"), str(input_file)],
                capture_output=True, text=True, timeout=120, cwd=str(BASE_DIR)
            )
            if result.returncode != 0:
                raise HTTPException(status_code=500, detail={"error": "parse_cdf.py failed", "stage": "parse", "stderr": result.stderr[-500:]})
            stages_completed.append("parse")
            WATCHER_STATUS["stages_completed"] = ["parse"]
            WATCHER_STATUS["stage"] = "deconvolving"
        except subprocess.TimeoutExpired:
            raise HTTPException(status_code=504, detail="parse_cdf.py timed out")

        # Deconvolve
        parse_output = PROC_DIR / f"{sample_id}.json"
        if parse_output.exists():
            try:
                result = subprocess.run(
                    [PYTHON, str(SCRIPTS_DIR / "gnn_deconv.py"), str(parse_output)],
                    capture_output=True, text=True, timeout=120, cwd=str(BASE_DIR)
                )
                if result.returncode == 0:
                    stages_completed.append("deconv")
                    WATCHER_STATUS["stages_completed"] = ["parse", "deconv"]
                    WATCHER_STATUS["stage"] = "matching"
            except subprocess.TimeoutExpired:
                pass  # Continue to match even if deconv times out

        # Match
        match_input = PROC_DIR / f"{sample_id}_deconvolved.json"
        if not match_input.exists():
            match_input = parse_output

        if match_input.exists():
            try:
                result = subprocess.run(
                    [PYTHON, str(SCRIPTS_DIR / "spectral_match.py"), str(match_input)],
                    capture_output=True, text=True, timeout=300, cwd=str(BASE_DIR)
                )
                if result.returncode == 0:
                    stages_completed.append("match")
                    WATCHER_STATUS["stages_completed"] = ["parse", "deconv", "match"]
                    WATCHER_STATUS["stage"] = "enriching"
            except subprocess.TimeoutExpired:
                print("[FULL PIPELINE] Matching timed out after 300s.")
                pass

        # Enrich
        matched_output = PROC_DIR / f"{sample_id}_matched.json"
        if matched_output.exists():
            try:
                matched_data = json.loads(matched_output.read_text())
                peaks = matched_data.get("peaks", [])
                if peaks:
                    enrich_req = EnrichRequest(
                        sample_id=sample_id,
                        peaks=peaks,
                        context="GC-MS chromatography, chemical analysis"
                    )
                    run_enrich(enrich_req)
                    stages_completed.append("enrich")
                    WATCHER_STATUS["stages_completed"] = ["parse", "deconv", "match", "enrich"]
                    WATCHER_STATUS["stage"] = "generating_excel"
            except Exception as enrich_err:
                print(f"[FULL PIPELINE] AI enrichment failed: {enrich_err}")

        # Automatically generate Excel spreadsheet output in processed_results/
        excel_report_path = RESULTS_DIR / f"{sample_id}_results.xlsx"
        try:
            print(f"[FULL PIPELINE] Generating final Excel report for {sample_id}...")
            create_excel_report(sample_id)
        except Exception as excel_err:
            print(f"[FULL PIPELINE] Excel report generation failed: {excel_err}")

        # FAIR Storage Layer (Zarr & LaminDB)
        try:
            import sys
            sys.path.append(str(SCRIPTS_DIR))
            import data_store
            
            # 1. Zarr write
            zarr_path = data_store.save_run_to_zarr(sample_id, str(input_file))
            # 2. LaminDB registration
            db_res = data_store.register_in_lamindb(sample_id, str(excel_report_path))
            print(f"[FULL PIPELINE STORE] Registered in Zarr: {zarr_path} | LaminDB: {db_res}")
        except Exception as store_err:
            print(f"[FULL PIPELINE STORE ERROR] Storage/LaminDB failed: {store_err}")

        # Load final result
        final_file = PROC_DIR / f"{sample_id}_enriched.json"
        if not final_file.exists():
            final_file = PROC_DIR / f"{sample_id}_matched.json"
        if not final_file.exists():
            final_file = PROC_DIR / f"{sample_id}_deconvolved.json"
        if not final_file.exists():
            final_file = parse_output

        final_data = {}
        if final_file.exists():
            final_data = json.loads(final_file.read_text())

        # Retrieve match count from matched file if it exists
        match_count = 0
        matched_file = PROC_DIR / f"{sample_id}_matched.json"
        if matched_file.exists():
            try:
                match_count = len(json.loads(matched_file.read_text()).get("matches", []))
            except Exception:
                pass

        # Complete status
        WATCHER_STATUS["active"] = False
        WATCHER_STATUS["stage"] = "complete"

        return {
            "status": "ok",
            "sample_id": sample_id,
            "stages_completed": stages_completed,
            "peak_count": len(final_data.get("peaks", [])),
            "match_count": match_count,
        }
    except Exception as run_err:
        WATCHER_STATUS["active"] = False
        WATCHER_STATUS["stage"] = "failed"
        WATCHER_STATUS["last_error"] = str(run_err)
        raise run_err


# ── Autonomous Folder Watcher ────────────────────────────────────────

def run_autonomous_pipeline(file_path: Path, sample_id: str, currently_processing: set):
    """Execute chromatography pipeline stages sequentially: parse -> deconv -> match -> enrich.
    Upon completion, automatically creates the Excel result spreadsheet in RESULTS_DIR."""
    try:
        WATCHER_STATUS["active"] = True
        WATCHER_STATUS["current_file"] = file_path.name
        WATCHER_STATUS["stage"] = "parsing"
        WATCHER_STATUS["stages_completed"] = []
        WATCHER_STATUS["last_error"] = None

        print(f"[AUTONOMOUS RUN] Step 1/4: Ingesting & parsing raw data: {file_path.name}")
        # Run parse
        parse_res = subprocess.run(
            [PYTHON, str(SCRIPTS_DIR / "parse_cdf.py"), str(file_path)],
            capture_output=True, text=True, timeout=120, cwd=str(BASE_DIR)
        )
        if parse_res.returncode != 0:
            err_msg = f"Parsing failed for {file_path.name}:\n{parse_res.stderr[-300:]}"
            print(f"[AUTONOMOUS RUN ERROR] {err_msg}")
            WATCHER_STATUS["stage"] = "failed"
            WATCHER_STATUS["last_error"] = err_msg
            return
            
        parse_output = PROC_DIR / f"{sample_id}.json"
        if not parse_output.exists():
            err_msg = f"Parse output file not found: {parse_output}"
            print(f"[AUTONOMOUS RUN ERROR] {err_msg}")
            WATCHER_STATUS["stage"] = "failed"
            WATCHER_STATUS["last_error"] = err_msg
            return
            
        WATCHER_STATUS["stages_completed"].append("parse")
        WATCHER_STATUS["stage"] = "deconvolving"

        print(f"[AUTONOMOUS RUN] Step 2/4: GNN Deconvolution on {sample_id}")
        # Run deconv
        subprocess.run(
            [PYTHON, str(SCRIPTS_DIR / "gnn_deconv.py"), str(parse_output)],
            capture_output=True, text=True, timeout=120, cwd=str(BASE_DIR)
        )
        
        WATCHER_STATUS["stages_completed"].append("deconv")
        WATCHER_STATUS["stage"] = "matching"

        # Run match
        match_input = PROC_DIR / f"{sample_id}_deconvolved.json"
        if not match_input.exists():
            match_input = parse_output
            
        print(f"[AUTONOMOUS RUN] Step 3/4: Spectral Similarity Matching (matchms) on {sample_id}")
        subprocess.run(
            [PYTHON, str(SCRIPTS_DIR / "spectral_match.py"), str(match_input)],
            capture_output=True, text=True, timeout=300, cwd=str(BASE_DIR)
        )
        
        WATCHER_STATUS["stages_completed"].append("match")
        WATCHER_STATUS["stage"] = "enriching"

        # Run enrich
        matched_output = PROC_DIR / f"{sample_id}_matched.json"
        if matched_output.exists():
            print(f"[AUTONOMOUS RUN] Step 4/4: T2 AI Enrichment on {sample_id}")
            try:
                matched_data = json.loads(matched_output.read_text())
                peaks = matched_data.get("peaks", [])
                if peaks:
                    enrich_req = EnrichRequest(
                        sample_id=sample_id,
                        peaks=peaks,
                        context="GC-MS chromatography, chemical analysis, autonomous trigger"
                    )
                    run_enrich(enrich_req)
                    
                    WATCHER_STATUS["stages_completed"].append("enrich")
                    WATCHER_STATUS["stage"] = "generating_excel"

                    # Generate the Excel report autonomously!
                    print(f"[AUTONOMOUS RUN] Generating final Excel report for {sample_id}...")
                    create_excel_report(sample_id)

                    # FAIR Storage Layer (Zarr & LaminDB)
                    try:
                        import sys
                        sys.path.append(str(SCRIPTS_DIR))
                        import data_store
                        
                        # 1. Zarr write
                        zarr_path = data_store.save_run_to_zarr(sample_id, str(file_path))
                        # 2. LaminDB registration
                        excel_report_path = RESULTS_DIR / f"{sample_id}_results.xlsx"
                        db_res = data_store.register_in_lamindb(sample_id, str(excel_report_path))
                        print(f"[AUTONOMOUS STORE] Registered in Zarr: {zarr_path} | LaminDB: {db_res}")
                    except Exception as store_err:
                        print(f"[AUTONOMOUS STORE ERROR] Storage/LaminDB failed: {store_err}")

                    print(f"[AUTONOMOUS RUN] Full pipeline completed successfully for {sample_id}!")
                    WATCHER_STATUS["stage"] = "complete"
            except Exception as enrich_err:
                print(f"[AUTONOMOUS RUN ERROR] Enrichment failed: {enrich_err}")
                WATCHER_STATUS["stage"] = "failed"
                WATCHER_STATUS["last_error"] = f"Enrichment failed: {enrich_err}"
        else:
            print(f"[AUTONOMOUS RUN ERROR] Match output file not found for {sample_id}")
            WATCHER_STATUS["stage"] = "failed"
            WATCHER_STATUS["last_error"] = "Match output file not found"
            
    except Exception as pipe_err:
        print(f"[AUTONOMOUS RUN ERROR] Pipeline execution exception for {sample_id}: {pipe_err}")
        WATCHER_STATUS["stage"] = "failed"
        WATCHER_STATUS["last_error"] = str(pipe_err)
    finally:
        currently_processing.discard(sample_id)
        # Sleep for 1 second just to let final UI state sync, but do NOT clear the completed status
        time.sleep(1)
        if WATCHER_STATUS["current_file"] == file_path.name:
            WATCHER_STATUS["active"] = False


def watch_raw_data_folder():
    """Background directory poll watcher for autonomous chromatography file ingestion.
    Checks raw_data/ every 3 seconds for new .cdf or .mzML files, waits for file write completion,
    then automatically executes the full parse -> deconv -> match -> enrich pipeline."""
    print(f"[AUTONOMOUS TRIGGER] Folder watcher started on {RAW_DIR}")
    
    file_sizes = {}
    currently_processing = set()
    
    while True:
        try:
            raw_files = (list(RAW_DIR.glob("*.cdf")) + list(RAW_DIR.glob("*.mzML")) + list(RAW_DIR.glob("*.CDF")) +
                         list(RAW_DIR.glob("*.xms")) + list(RAW_DIR.glob("*.XMS")) +
                         list(RAW_DIR.glob("*.sms")) + list(RAW_DIR.glob("*.SMS")))
            
            for path in raw_files:
                sample_id = path.stem
                
                # Check if already processed (exists in cache or results)
                enriched_file = PROC_DIR / f"{sample_id}_enriched.json"
                matched_file = PROC_DIR / f"{sample_id}_matched.json"
                parsed_file = PROC_DIR / f"{sample_id}.json"
                excel_file = RESULTS_DIR / f"{sample_id}_results.xlsx"
                
                # If the final Excel results file already exists, we consider it processed
                if excel_file.exists():
                    continue
                
                # If currently processing, skip
                if sample_id in currently_processing:
                    continue
                
                # File size stability check (to ensure GC-MS/HPLC finished writing)
                curr_size = path.stat().st_size
                prev_size = file_sizes.get(path)
                
                if prev_size is None or curr_size != prev_size:
                    # File size is changing or just detected, record size and wait
                    file_sizes[path] = curr_size
                    print(f"[AUTONOMOUS TRIGGER] Detected file write in progress: {path.name} (size: {curr_size} bytes)")
                    continue
                
                # Size has not changed since last check (3s ago), safe to process
                print(f"[AUTONOMOUS TRIGGER] File size stabilized: {path.name}. Starting autonomous run...")
                currently_processing.add(sample_id)
                
                # Initialize status
                WATCHER_STATUS["active"] = True
                WATCHER_STATUS["current_file"] = path.name
                WATCHER_STATUS["stage"] = "starting"
                WATCHER_STATUS["stages_completed"] = []
                WATCHER_STATUS["last_error"] = None

                # Run full pipeline in a separate thread
                pipeline_thread = threading.Thread(
                    target=run_autonomous_pipeline,
                    args=(path, sample_id, currently_processing),
                    daemon=True
                )
                pipeline_thread.start()
                
        except Exception as scan_err:
            print(f"[AUTONOMOUS TRIGGER ERROR] Error scanning folder: {scan_err}")
            
        time.sleep(3)


def start_folder_watcher():
    """Start the background directory watcher thread."""
    watcher_thread = threading.Thread(target=watch_raw_data_folder, daemon=True)
    watcher_thread.start()





if __name__ == "__main__":
    print("=" * 55)
    print("  CHROMA-AGENT-ALPHA Pipeline Server")
    print("  http://localhost:8001")
    print("  Docs: http://localhost:8001/docs")
    print("=" * 55)
    uvicorn.run(app, host="0.0.0.0", port=8001, log_level="info")
